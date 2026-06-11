"""Export extraction results to an .xlsx workbook.

Layout:
  - "Data" sheet: one row per document, one column per scalar field.
  - One extra sheet per table field, with the table's columns plus a "_file"
    column linking each row back to its source document.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 60)


def export_results(template: dict, results: list[dict], exports_dir: Path) -> str:
    """results: [{"file": name, "fields": {...}}]. Returns the saved file path."""
    fields = template["fields"]
    scalar_fields = [f for f in fields if f["type"] != "table"]
    table_fields = [f for f in fields if f["type"] == "table"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["_file"] + [f["name"] for f in scalar_fields]
    ws.append(headers)
    for res in results:
        row = [res.get("file", "")]
        for f in scalar_fields:
            row.append(res["fields"].get(f["name"]))
        ws.append(row)
    _style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    _autosize(ws)

    for tf in table_fields:
        cols = [c["name"] for c in tf.get("columns", [])]
        # Sheet titles are capped at 31 chars and can't contain some symbols.
        title = tf["name"][:28] or "table"
        tws = wb.create_sheet(title=title)
        tws.append(["_file"] + cols)
        for res in results:
            rows = res["fields"].get(tf["name"]) or []
            for r in rows:
                tws.append([res.get("file", "")] + [r.get(c) for c in cols])
        _style_header(tws, len(cols) + 1)
        tws.freeze_panes = "A2"
        _autosize(tws)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe = template.get("name", "export").replace(" ", "_")[:40]
    fname = f"{safe}_{stamp}_{uuid.uuid4().hex[:6]}.xlsx"
    path = exports_dir / fname
    wb.save(path)
    return str(path)
